#!/usr/bin/env python3
"""
Convert the official Polish driving license question database from XLSX to CSV.

Usage:
    python scripts/convert-questions.py <input.xlsx> [output.csv]

Download the official database from:
    https://www.gov.pl/attachment/921380de-3ac3-480d-b802-30aa03a67462

The script auto-detects the column layout used in the official gov.pl spreadsheet
and outputs a CSV file in the format expected by the quiz application:
    id,question,answer_a,answer_b,answer_c,correct,media,points,category
"""

import sys
import csv
import re
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Please install openpyxl: pip install openpyxl")
    sys.exit(1)


# Mapping from possible Polish column names to internal names
COLUMN_ALIASES = {
    "id": ["id", "numer pytania", "numer", "lp"],
    "question": ["pytanie", "treść pytania", "tresc pytania", "question"],
    "answer_a": ["odpowiedź a", "odpowiedz a", "odp. a", "a"],
    "answer_b": ["odpowiedź b", "odpowiedz b", "odp. b", "b"],
    "answer_c": ["odpowiedź c", "odpowiedz c", "odp. c", "c"],
    "correct": [
        "prawidłowa odpowiedź",
        "prawidlowa odpowiedz",
        "odpowiedź prawidłowa",
        "poprawna",
        "correct",
    ],
    "media": ["plik multimedialny", "multimedia", "media", "plik", "zdjęcie"],
    "points": ["pkt", "punkty", "waga", "points", "wartość punktowa"],
    "category": ["kategorie", "kategoria", "category", "kat."],
}


def normalize(text: str) -> str:
    return text.strip().lower()


def detect_columns(header_row: list) -> dict:
    """Return a mapping from internal name to column index."""
    header = [normalize(str(cell.value or "")) for cell in header_row]
    mapping = {}
    for internal_name, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            try:
                idx = next(i for i, h in enumerate(header) if alias in h)
                mapping[internal_name] = idx
                break
            except StopIteration:
                continue
    return mapping


def convert_xlsx(input_path: str, output_path: str) -> None:
    wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)

    # Try to find the sheet with questions (usually first or named "Pytania")
    sheet = None
    for name in wb.sheetnames:
        if any(k in name.lower() for k in ["pytan", "question", "baza"]):
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb.active

    rows = list(sheet.iter_rows())
    if not rows:
        print("ERROR: No rows found in spreadsheet.")
        sys.exit(1)

    # Find header row (first row with ≥4 non-empty cells)
    header_row_idx = 0
    for i, row in enumerate(rows):
        non_empty = sum(1 for cell in row if cell.value is not None)
        if non_empty >= 4:
            header_row_idx = i
            break

    col_map = detect_columns(rows[header_row_idx])
    print(f"Detected columns: {col_map}")

    required = ["question", "answer_a", "answer_b", "answer_c", "correct"]
    missing = [r for r in required if r not in col_map]
    if missing:
        print(f"WARNING: Could not detect columns: {missing}")
        print("Available headers:", [str(c.value) for c in rows[header_row_idx]])

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, quoting=csv.QUOTE_ALL)
        writer.writerow(
            ["id", "question", "answer_a", "answer_b", "answer_c", "correct",
             "media", "points", "category"]
        )

        converted = 0
        for row in rows[header_row_idx + 1 :]:
            vals = [str(cell.value).strip() if cell.value is not None else "" for cell in row]
            if not vals or all(v == "" for v in vals):
                continue

            def get(key: str, default="") -> str:
                idx = col_map.get(key)
                if idx is None or idx >= len(vals):
                    return default
                return vals[idx].strip()

            q_id = get("id", str(converted + 1))
            question = get("question")
            answer_a = get("answer_a")
            answer_b = get("answer_b")
            answer_c = get("answer_c")
            correct_raw = get("correct", "").upper()

            # Normalize correct answer: accept "A", "B", "C" or numeric "1","2","3"
            correct = correct_raw
            if correct_raw in ("1",):
                correct = "A"
            elif correct_raw in ("2",):
                correct = "B"
            elif correct_raw in ("3",):
                correct = "C"
            correct = re.sub(r"[^ABC]", "", correct)
            if not correct:
                continue

            media = get("media")
            points_raw = get("points", "1")
            points = "3" if points_raw.strip() in ("3", "3.0") else "1"
            category = get("category", "B")

            if not question or not answer_a or not answer_b or not answer_c:
                continue

            writer.writerow(
                [q_id, question, answer_a, answer_b, answer_c, correct,
                 media, points, category]
            )
            converted += 1

    print(f"Converted {converted} questions → {output_path}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else "data/questions.csv"

    convert_xlsx(input_file, output_file)
